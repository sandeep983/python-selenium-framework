from os import set_handle_inheritable
import openpyxl

book = openpyxl.load_workbook("C:\\Users\\sande\\Desktop\\Programs\\Projects\\python-selenium-framework\\testData\\xldata.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)

for r in range(1, sheet.max_row+1):
    if(sheet.cell(row=r, column=1).value == 'TestCase3'):
        for c in range(1, sheet.max_column+1):
            print(sheet.cell(row=r, column=c).value, end=" ")